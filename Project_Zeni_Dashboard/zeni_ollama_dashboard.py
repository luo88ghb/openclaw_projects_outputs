#!/usr/bin/env python3
"""
Zeni Ollama Dashboard - Daily Token Usage Report
監控 Ollama Cloud 免費模型每日 Token 流量與費用計算
作者: Zeni (傑尼)
版本: 1.0.0
日期: 2026-03-20
"""

import json
import os
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
import sys

try:
    import requests
except ImportError:
    import urllib.request
    def requests_get(url, timeout=None):
        with urllib.request.urlopen(url, timeout=timeout) as response:
            return type('Response', (), {'status_code': response.status, 'json': lambda: json.loads(response.read())})()
    requests = type('requests', (), {'get': requests_get})()

# Configuration
DB_PATH = Path.home() / ".openclaw" / "memory" / "token_usage.db"
REPORTS_DIR = Path.home() / ".openclaw" / "workspace" / "Project_Zeni_Dashboard" / "reports"
MONTHLY_REPORTS_DIR = REPORTS_DIR / "monthly"

# Model Pricing (USD per 1K tokens) - 根據官方調用費
MODEL_PRICING = {
    # Ollama Cloud Free Models (免費額度內)
    "gpt-oss:120b-cloud": {
        "input": 0.0000,   # 免費
        "output": 0.0000,  # 免費
        "free_limit": 10000,  # 7天免費額度
        "description": "GPT-OSS 120B (Cloud Free)"
    },
    # Local Ollama Models (免費)
    "ollama/qwen359b:latest": {
        "input": 0.0000,
        "output": 0.0000,
        "description": "Qwen3.5 9B (Local)"
    },
    "ollama/gemma3-tools:12b": {
        "input": 0.0000,
        "output": 0.0000,
        "description": "Gemma3 Tools 12B (Local)"
    },
    "ollama/qwen3.5:9b": {
        "input": 0.0000,
        "output": 0.0000,
        "description": "Qwen3.5 9B (Local)"
    },
    "ollama/qwen3-vl:latest": {
        "input": 0.0000,
        "output": 0.0000,
        "description": "Qwen3-VL (Local Vision)"
    },
    "ollama/DeepSeek-Janus-Pro-7B": {
        "input": 0.0000,
        "output": 0.0000,
        "description": "DeepSeek Janus Pro 7B (Local)"
    },
    # Cloud Paid Models (for reference)
    "kimi-k2.5:cloud": {
        "input": 0.0020,
        "output": 0.0060,
        "description": "Kimi K2.5 (Cloud)"
    },
    "minimax-m2.5:cloud": {
        "input": 0.0015,
        "output": 0.0045,
        "description": "MiniMax M2.5 (Cloud)"
    }
}

# Ollama Cloud Free Limit
FREE_TIER_LIMIT = 10000  # tokens per 7 days


def init_database():
    """Initialize SQLite database for token tracking"""
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Daily usage table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS daily_usage (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            model TEXT NOT NULL,
            input_tokens INTEGER DEFAULT 0,
            output_tokens INTEGER DEFAULT 0,
            total_tokens INTEGER DEFAULT 0,
            estimated_cost_usd REAL DEFAULT 0.0,
            estimated_cost_twd REAL DEFAULT 0.0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Create index for faster queries
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_daily_usage_date 
        ON daily_usage(date)
    ''')
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_daily_usage_model 
        ON daily_usage(model)
    ''')
    
    conn.commit()
    conn.close()
    print(f"✅ Database initialized at {DB_PATH}")


def get_openclaw_token_usage():
    """Get token usage from OpenClaw session logs"""
    # Try to read from OpenClaw's internal tracking
    # This is a placeholder - actual implementation would parse OpenClaw logs
    
    usage_data = {
        "gpt-oss:120b-cloud": {"input": 0, "output": 0},
        "ollama/qwen359b:latest": {"input": 0, "output": 0},
        "ollama/qwen3.5:9b": {"input": 0, "output": 0},
        "kimi-k2.5:cloud": {"input": 0, "output": 0},
        "minimax-m2.5:cloud": {"input": 0, "output": 0}
    }
    
    # Read from OpenClaw memory files
    memory_dir = Path.home() / ".openclaw" / "memory"
    today = datetime.now().strftime("%Y-%m-%d")
    
    # Look for today's memory files
    for mem_file in memory_dir.glob("*.md"):
        if today in mem_file.name or "MEMORY" in mem_file.name:
            try:
                content = mem_file.read_text(encoding='utf-8')
                # Parse token usage from memory files
                if "輸入" in content and "輸出" in content:
                    # Extract model usage from the file
                    lines = content.split('\n')
                    for line in lines:
                        if "執行模型:" in line:
                            model = line.split("執行模型:")[1].strip() if "執行模型:" in line else ""
                            if model and model in usage_data:
                                # Extract token counts from nearby lines
                                pass
            except Exception as e:
                print(f"Warning: Could not read {mem_file}: {e}")
    
    return usage_data


def get_ollama_local_status():
    """Get Ollama local server status"""
    try:
        response = requests.get("http://localhost:11434/api/tags", timeout=5)
        if response.status_code == 200:
            models = response.json().get("models", [])
            return {
                "status": "✅ Running",
                "models_loaded": len(models),
                "models": [m.get("name", "unknown") for m in models]
            }
    except Exception as e:
        pass
    
    return {
        "status": "❌ Offline",
        "models_loaded": 0,
        "models": []
    }


def get_system_resource_usage():
    """Get system resource usage"""
    try:
        import psutil
        
        cpu_percent = psutil.cpu_percent(interval=1)
        memory = psutil.virtual_memory()
        disk = psutil.disk_usage('/')
        
        return {
            "cpu_percent": cpu_percent,
            "memory_used_gb": memory.used / (1024**3),
            "memory_total_gb": memory.total / (1024**3),
            "memory_percent": memory.percent,
            "disk_used_gb": disk.used / (1024**3),
            "disk_total_gb": disk.total / (1024**3),
            "disk_percent": disk.percent
        }
    except ImportError:
        return {
            "cpu_percent": 0,
            "memory_used_gb": 0,
            "memory_total_gb": 0,
            "memory_percent": 0,
            "disk_used_gb": 0,
            "disk_total_gb": 0,
            "disk_percent": 0
        }


def calculate_cost(model, input_tokens, output_tokens):
    """Calculate estimated cost for token usage"""
    pricing = MODEL_PRICING.get(model, {"input": 0, "output": 0})
    
    input_cost = (input_tokens / 1000) * pricing["input"]
    output_cost = (output_tokens / 1000) * pricing["output"]
    total_usd = input_cost + output_cost
    
    # Convert to TWD (approximate rate: 1 USD = 32 TWD)
    total_twd = total_usd * 32
    
    return {
        "usd": total_usd,
        "twd": total_twd
    }


def get_monthly_stats(year_month):
    """Get monthly token usage statistics"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT 
            model,
            SUM(input_tokens) as total_input,
            SUM(output_tokens) as total_output,
            SUM(total_tokens) as total_tokens,
            SUM(estimated_cost_usd) as total_cost_usd,
            SUM(estimated_cost_twd) as total_cost_twd
        FROM daily_usage
        WHERE date LIKE ?
        GROUP BY model
    ''', (f"{year_month}%",))
    
    results = cursor.fetchall()
    conn.close()
    
    return results


def generate_daily_report():
    """Generate daily token usage report"""
    today = datetime.now()
    today_str = today.strftime("%Y-%m-%d")
    
    # Get usage data
    token_usage = get_openclaw_token_usage()
    ollama_status = get_ollama_local_status()
    system_resources = get_system_resource_usage()
    
    # Calculate daily totals
    daily_totals = {
        "input": 0,
        "output": 0,
        "cost_usd": 0.0,
        "cost_twd": 0.0
    }
    
    model_breakdown = []
    
    for model, tokens in token_usage.items():
        cost = calculate_cost(model, tokens["input"], tokens["output"])
        
        daily_totals["input"] += tokens["input"]
        daily_totals["output"] += tokens["output"]
        daily_totals["cost_usd"] += cost["usd"]
        daily_totals["cost_twd"] += cost["twd"]
        
        model_breakdown.append({
            "model": model,
            "description": MODEL_PRICING.get(model, {}).get("description", model),
            "input_tokens": tokens["input"],
            "output_tokens": tokens["output"],
            "total_tokens": tokens["input"] + tokens["output"],
            "cost_usd": cost["usd"],
            "cost_twd": cost["twd"]
        })
    
    # Get monthly stats
    year_month = today.strftime("%Y-%m")
    monthly_stats = get_monthly_stats(year_month)
    
    monthly_total = {
        "tokens": 0,
        "cost_usd": 0.0,
        "cost_twd": 0.0
    }
    
    for row in monthly_stats:
        monthly_total["tokens"] += row[3]  # total_tokens
        monthly_total["cost_usd"] += row[4]  # total_cost_usd
        monthly_total["cost_twd"] += row[5]  # total_cost_twd
    
    # Calculate free tier remaining
    free_used = sum([
        tokens["input"] + tokens["output"] 
        for model, tokens in token_usage.items() 
        if "cloud" in model and MODEL_PRICING.get(model, {}).get("free_limit", 0) > 0
    ])
    free_remaining = max(0, FREE_TIER_LIMIT - free_used)
    
    # Generate report
    report = {
        "date": today_str,
        "daily": {
            "total_input_tokens": daily_totals["input"],
            "total_output_tokens": daily_totals["output"],
            "total_tokens": daily_totals["input"] + daily_totals["output"],
            "estimated_cost_usd": daily_totals["cost_usd"],
            "estimated_cost_twd": daily_totals["cost_twd"]
        },
        "monthly": {
            "total_tokens": monthly_total["tokens"] + daily_totals["input"] + daily_totals["output"],
            "total_cost_usd": monthly_total["cost_usd"] + daily_totals["cost_usd"],
            "total_cost_twd": monthly_total["cost_twd"] + daily_totals["cost_twd"]
        },
        "free_tier": {
            "limit": FREE_TIER_LIMIT,
            "used": free_used,
            "remaining": free_remaining,
            "reset_date": (today + timedelta(days=7)).strftime("%Y-%m-%d")
        },
        "models": model_breakdown,
        "ollama_status": ollama_status,
        "system_resources": system_resources
    }
    
    return report


def save_to_database(report):
    """Save daily report to SQLite database"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    for model_data in report["models"]:
        cursor.execute('''
            INSERT OR REPLACE INTO daily_usage 
            (date, model, input_tokens, output_tokens, total_tokens, 
             estimated_cost_usd, estimated_cost_twd)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            report["date"],
            model_data["model"],
            model_data["input_tokens"],
            model_data["output_tokens"],
            model_data["total_tokens"],
            model_data["cost_usd"],
            model_data["cost_twd"]
        ))
    
    conn.commit()
    conn.close()


def generate_monthly_excel(year_month=None):
    """Generate monthly Excel report"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("Warning: openpyxl not installed. Install with: pip install openpyxl")
        return None
    
    if year_month is None:
        year_month = datetime.now().strftime("%Y-%m")
    
    monthly_stats = get_monthly_stats(year_month)
    
    if not monthly_stats:
        print(f"No data found for {year_month}")
        return None
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Report"
    
    # Headers
    headers = ["Model", "Input Tokens", "Output Tokens", "Total Tokens", "Cost (USD)", "Cost (TWD)"]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    total_input = total_output = total_tokens = 0
    total_cost_usd = total_cost_twd = 0.0
    
    for row in monthly_stats:
        model = row[0]
        input_t = row[1]
        output_t = row[2]
        total_t = row[3]
        cost_usd = row[4]
        cost_twd = row[5]
        
        ws.append([
            MODEL_PRICING.get(model, {}).get("description", model),
            input_t,
            output_t,
            total_t,
            round(cost_usd, 4),
            round(cost_twd, 2)
        ])
        
        total_input += input_t
        total_output += output_t
        total_tokens += total_t
        total_cost_usd += cost_usd
        total_cost_twd += cost_twd
    
    # Totals row
    ws.append([])
    ws.append([
        "TOTAL",
        total_input,
        total_output,
        total_tokens,
        round(total_cost_usd, 4),
        round(total_cost_twd, 2)
    ])
    
    # Style totals
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    MONTHLY_REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    excel_path = MONTHLY_REPORTS_DIR / f"Zeni_Ollama_Dashboard_{year_month}.xlsx"
    wb.save(excel_path)
    
    print(f"✅ Monthly Excel report saved: {excel_path}")
    return excel_path


def format_telegram_message(report):
    """Format report for Telegram notification"""
    today = report["date"]
    daily = report["daily"]
    monthly = report["monthly"]
    free_tier = report["free_tier"]
    ollama = report["ollama_status"]
    sys_res = report["system_resources"]
    
    # Calculate free tier percentage
    free_percent = (free_tier["used"] / free_tier["limit"]) * 100 if free_tier["limit"] > 0 else 0
    
    message = f"""📊 **Zeni Ollama Dashboard - Daily Report**
📅 {today} 23:00 Taipei Time

💰 **今日使用統計**
├ Input Tokens: {daily['total_input_tokens']:,}
├ Output Tokens: {daily['total_output_tokens']:,}
├ Total Tokens: {daily['total_tokens']:,}
├ Est. Cost: ${daily['estimated_cost_usd']:.4f} USD
└ Est. Cost: ${daily['estimated_cost_twd']:.2f} TWD

📈 **本月累計 (至今日)**
├ Total Tokens: {monthly['total_tokens']:,}
├ Est. Cost: ${monthly['total_cost_usd']:.4f} USD
└ Est. Cost: ${monthly['total_cost_twd']:.2f} TWD

🎁 **免費額度狀態 (Ollama Cloud)**
├ Limit: {free_tier['limit']:,} tokens / 7 days
├ Used: {free_tier['used']:,} ({free_percent:.1f}%)
├ Remaining: {free_tier['remaining']:,}
└ Reset: {free_tier['reset_date']}

🖥️ **Ollama 本地狀態**
├ Status: {ollama['status']}
├ Models Loaded: {ollama['models_loaded']}
└ Models: {', '.join(ollama['models'][:3]) if ollama['models'] else 'N/A'}

⚙️ **系統資源**
├ CPU: {sys_res['cpu_percent']:.1f}%
├ Memory: {sys_res['memory_percent']:.1f}% ({sys_res['memory_used_gb']:.1f}/{sys_res['memory_total_gb']:.1f} GB)
└ Disk: {sys_res['disk_percent']:.1f}% ({sys_res['disk_used_gb']:.1f}/{sys_res['disk_total_gb']:.1f} GB)

🐢 Zeni 的靈魂記憶！每日自動彙整報告
"""
    return message


def main():
    """Main function"""
    # Set UTF-8 encoding for Windows
    import sys
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
    
    print("Zeni Ollama Dashboard - Starting Daily Report...")
    
    # Initialize database
    init_database()
    
    # Generate daily report
    report = generate_daily_report()
    
    # Save to database
    save_to_database(report)
    
    # Check if it's month-end (last day of month)
    today = datetime.now()
    next_day = today + timedelta(days=1)
    is_month_end = next_day.day == 1
    
    excel_path = None
    if is_month_end:
        print("📅 Month end detected - generating Excel report...")
        last_month = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
        excel_path = generate_monthly_excel(last_month)
    
    # Output for Telegram - Clean format without extra text
    telegram_message = format_telegram_message(report)
    
    # Save report to file for reference
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    report_file = REPORTS_DIR / f"daily_report_{report['date']}.json"
    with open(report_file, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    
    # Print ONLY the telegram message - this will be captured by OpenClaw
    print(telegram_message)
    
    return telegram_message


if __name__ == "__main__":
    try:
        result = main()
        print(result)
    except Exception as e:
        error_msg = f"Zeni Ollama Dashboard Error: {str(e)}"
        print(error_msg)
        sys.exit(1)
